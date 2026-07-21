"""
apollo_core.py

Reusable Apollo.io enrichment logic, shared by the CLI script and the
Streamlit app. No argparse/CLI code here — just functions.
"""

import time
import requests

API_BASE = "https://api.apollo.io/api/v1"

TARGET_TITLES_DEFAULT = [
    "Purchasing Manager",
    "Procurement Manager",
    "Head of Purchasing",
    "Head of Procurement",
    "Buyer",
    "Category Manager",
    "Sourcing Manager",
    "Chief Procurement Officer",
]

REQUEST_DELAY_SECONDS = 1.0


class ApolloError(Exception):
    pass


def apollo_post(endpoint, api_key, payload, retries=1):
    url = f"{API_BASE}/{endpoint}"
    headers = {"Content-Type": "application/json", "X-Api-Key": api_key}
    resp = requests.post(url, headers=headers, json=payload, timeout=30)
    if resp.status_code == 401:
        raise ApolloError("401 Unauthorized — check the API key / plan access.")
    if resp.status_code == 429 and retries > 0:
        time.sleep(30)
        return apollo_post(endpoint, api_key, payload, retries=retries - 1)
    resp.raise_for_status()
    return resp.json()


def read_company_names(xlsx_file):
    """xlsx_file: path or file-like object with a 'Company name' column."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    header = [c.value for c in ws[1]]
    if "Company name" not in header:
        raise ValueError(f"Couldn't find a 'Company name' column. Columns found: {header}")
    name_col = header.index("Company name")

    names, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col]
        if name and str(name).strip():
            n = str(name).strip()
            if n not in seen:
                seen.add(n)
                names.append(n)
    return names


def match_company(name, api_key, target_titles):
    """Free: find the org + candidate people. Returns a dict, no email."""
    org_data = apollo_post(
        "mixed_companies/search", api_key, {"q_organization_name": name, "per_page": 1}
    )
    orgs = org_data.get("organizations", []) or org_data.get("accounts", [])
    if not orgs:
        return {"matched": False, "org_id": None, "people": []}

    org = orgs[0]
    org_id = org.get("id")

    people_data = apollo_post(
        "mixed_people/search",
        api_key,
        {"organization_ids": [org_id], "person_titles": target_titles, "per_page": 10},
    )
    people = people_data.get("people", [])

    return {
        "matched": True,
        "org_id": org_id,
        "org_name": org.get("name"),
        "org_domain": org.get("primary_domain") or org.get("website_url"),
        "people": [
            {
                "id": p.get("id"),
                "name": p.get("name"),
                "title": p.get("title"),
                "linkedin_url": p.get("linkedin_url"),
                "email_revealed": False,
                "email": None,
            }
            for p in people
        ],
    }


def reveal_person(person_id, api_key):
    """Costs 1 credit. Returns the email (or None)."""
    data = apollo_post(
        "people/match", api_key, {"id": person_id, "reveal_personal_emails": False}
    )
    return (data.get("person", {}) or {}).get("email")


def build_report_rows(companies_cache):
    """companies_cache: {company_name: match_company()-shaped dict}"""
    rows = []
    for name, info in companies_cache.items():
        if not info.get("matched"):
            rows.append([name, "", "", "", "", "", "No org match in Apollo"])
            continue
        people = info.get("people", [])
        if not people:
            rows.append([name, info.get("org_domain", ""), "", "", "", "",
                         "Org matched, no purchasing contact found"])
            continue
        for p in people:
            rows.append([
                name,
                info.get("org_domain", ""),
                p.get("name", ""),
                p.get("title", ""),
                p.get("email", ""),
                p.get("linkedin_url", ""),
                "Email revealed" if p.get("email_revealed") else "Found, email not revealed",
            ])
    return rows
